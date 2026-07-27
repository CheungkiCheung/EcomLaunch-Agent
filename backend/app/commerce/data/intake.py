"""Safe, deterministic intake for heterogeneous Commerce data files."""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import stat
import zipfile
from datetime import UTC, datetime
from enum import StrEnum
from pathlib import Path, PurePosixPath
from typing import Self

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, model_validator

from app.commerce.domain.ids import DatasetId, DataSourceId, WorkspaceId


class DataIntakeError(ValueError):
    """Raised when an uploaded bundle is unsafe, unsupported, or unreadable."""


class DatasetIntegrityError(DataIntakeError):
    """Raised when an already stored Dataset cannot be trusted or resumed."""


class FileFormat(StrEnum):
    CSV = "csv"
    JSON = "json"
    JSONL = "jsonl"
    XLSX = "xlsx"
    ZIP = "zip"


class IntakeModel(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)


class DataFileManifest(IntakeModel):
    id: DataSourceId = Field(default_factory=DataSourceId.new)
    original_name: str = Field(min_length=1)
    stored_relative_path: str = Field(min_length=1)
    format: FileFormat
    size_bytes: int = Field(ge=0)
    sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    encoding: str | None = Field(default=None, min_length=1)
    read_only: bool = True
    parent_source_id: DataSourceId | None = None
    archive_member: str | None = Field(default=None, min_length=1)


class TableManifest(IntakeModel):
    table_name: str = Field(min_length=1, pattern=r"^[a-z0-9]+(?:_[a-z0-9]+)*$")
    source_file_id: DataSourceId
    format: FileFormat
    sheet_name: str | None = Field(default=None, min_length=1)
    json_key: str | None = Field(default=None, min_length=1)
    archive_member: str | None = Field(default=None, min_length=1)


class DataBundleManifest(IntakeModel):
    schema_version: str = "1.0"
    dataset_id: DatasetId
    workspace_id: WorkspaceId
    created_at: datetime
    storage_relative_path: str = Field(min_length=1)
    files: tuple[DataFileManifest, ...] = Field(min_length=1)
    tables: tuple[TableManifest, ...] = Field(min_length=1)
    warnings: tuple[str, ...] = Field(default_factory=tuple)

    @model_validator(mode="after")
    def require_unique_ids_and_tables(self) -> Self:
        file_ids = [file.id for file in self.files]
        table_names = [table.table_name for table in self.tables]
        if len(file_ids) != len(set(file_ids)):
            raise ValueError("DataBundleManifest file IDs must be unique")
        if len(table_names) != len(set(table_names)):
            raise ValueError("DataBundleManifest table names must be unique")
        return self


class DataIntakeService:
    """Copy inputs into immutable storage and emit a verified table manifest."""

    MAX_ZIP_MEMBERS = 1_000
    MAX_ZIP_MEMBER_BYTES = 50 * 1024 * 1024
    MAX_ZIP_TOTAL_BYTES = 200 * 1024 * 1024
    MAX_ZIP_COMPRESSION_RATIO = 1_000

    def __init__(self, *, storage_root: Path):
        self._storage_root = storage_root

    def ingest(self, workspace_id: WorkspaceId, source_paths: tuple[Path, ...]) -> DataBundleManifest:
        if not source_paths:
            raise DataIntakeError("At least one input file is required")

        dataset_id = DatasetId.new()
        storage_relative_path = f"{workspace_id}/{dataset_id}"
        bundle_root = self._storage_root / storage_relative_path
        raw_root = bundle_root / "raw"
        expanded_root = bundle_root / "expanded"

        files: list[DataFileManifest] = []
        tables: list[TableManifest] = []
        warnings: list[str] = []
        table_names: set[str] = set()
        input_names: set[str] = set()

        try:
            raw_root.mkdir(parents=True, exist_ok=False)
            for source_path in source_paths:
                source_path = Path(source_path)
                if source_path.is_symlink():
                    raise DataIntakeError(f"Input symbolic links are not allowed: {source_path}")
                if not source_path.is_file():
                    raise DataIntakeError(f"Input file does not exist: {source_path}")
                if source_path.name in input_names:
                    raise DataIntakeError(f"Duplicate input filename: {source_path.name}")
                input_names.add(source_path.name)

                file_format = self._format_for_path(source_path)
                stored_path = raw_root / source_path.name
                shutil.copyfile(source_path, stored_path)
                self._make_read_only(stored_path)
                root_file = self._file_manifest(
                    stored_path,
                    bundle_root=bundle_root,
                    original_name=source_path.name,
                    file_format=file_format,
                )
                files.append(root_file)

                if file_format is FileFormat.ZIP:
                    member_files, member_tables, member_warnings = self._inspect_zip(
                        stored_path,
                        bundle_root=bundle_root,
                        expanded_root=expanded_root / str(root_file.id),
                        parent=root_file,
                        used_table_names=table_names,
                    )
                    files.extend(member_files)
                    tables.extend(member_tables)
                    warnings.extend(member_warnings)
                else:
                    tables.extend(
                        self._inspect_table_file(
                            stored_path,
                            source_file=root_file,
                            used_table_names=table_names,
                        )
                    )

            if not tables:
                raise DataIntakeError("Input bundle contains no supported tables")

            manifest = DataBundleManifest(
                dataset_id=dataset_id,
                workspace_id=workspace_id,
                created_at=datetime.now(UTC),
                storage_relative_path=storage_relative_path,
                files=tuple(files),
                tables=tuple(tables),
                warnings=tuple(warnings),
            )
            manifest_path = bundle_root / "manifest.json"
            manifest_path.write_text(manifest.model_dump_json(indent=2) + "\n", encoding="utf-8")
            self._make_read_only(manifest_path)
            return manifest
        except Exception:
            shutil.rmtree(bundle_root, ignore_errors=True)
            raise

    def verify_manifest(self, manifest: DataBundleManifest) -> None:
        """Verify that a persisted manifest still points to immutable source bytes."""

        expected_storage_path = f"{manifest.workspace_id}/{manifest.dataset_id}"
        if manifest.storage_relative_path != expected_storage_path:
            raise DatasetIntegrityError(f"Dataset manifest storage path does not match identity: {manifest.dataset_id}")

        bundle_root = self._storage_root / manifest.storage_relative_path
        storage_root = self._storage_root.resolve()
        resolved_bundle = bundle_root.resolve(strict=False)
        try:
            resolved_bundle.relative_to(storage_root)
        except ValueError as exc:
            raise DatasetIntegrityError(f"Dataset manifest storage path escapes storage root: {manifest.dataset_id}") from exc
        if not bundle_root.is_dir() or bundle_root.is_symlink():
            raise DatasetIntegrityError(f"Dataset storage directory is missing: {manifest.dataset_id}")

        file_ids = {file.id for file in manifest.files}
        for table in manifest.tables:
            if table.source_file_id not in file_ids:
                raise DatasetIntegrityError(f"Dataset table references a missing source file: {table.table_name}")

        for source_file in manifest.files:
            relative_path = PurePosixPath(source_file.stored_relative_path)
            if relative_path.is_absolute() or ".." in relative_path.parts or not relative_path.parts:
                raise DatasetIntegrityError(f"Dataset file path is unsafe: {source_file.stored_relative_path}")
            path = bundle_root.joinpath(*relative_path.parts)
            if path.is_symlink() or not path.is_file():
                raise DatasetIntegrityError(f"Dataset source file is missing: {source_file.original_name}")
            resolved_path = path.resolve(strict=False)
            try:
                resolved_path.relative_to(resolved_bundle)
            except ValueError as exc:
                raise DatasetIntegrityError(f"Dataset source file escapes its bundle: {source_file.original_name}") from exc
            if not source_file.read_only:
                raise DatasetIntegrityError(f"Dataset source file is not marked read-only: {source_file.original_name}")
            if path.stat().st_mode & (stat.S_IWUSR | stat.S_IWGRP | stat.S_IWOTH):
                raise DatasetIntegrityError(f"Dataset source file is writable: {source_file.original_name}")
            if path.stat().st_size != source_file.size_bytes:
                raise DatasetIntegrityError(f"Dataset source file size does not match its manifest: {source_file.original_name}")
            if self._sha256(path) != source_file.sha256:
                raise DatasetIntegrityError(f"Dataset source file hash does not match its manifest: {source_file.original_name}")

    @classmethod
    def _format_for_path(cls, path: Path) -> FileFormat:
        suffix = path.suffix.lower()
        mapping = {
            ".csv": FileFormat.CSV,
            ".json": FileFormat.JSON,
            ".jsonl": FileFormat.JSONL,
            ".ndjson": FileFormat.JSONL,
            ".xlsx": FileFormat.XLSX,
            ".zip": FileFormat.ZIP,
        }
        try:
            return mapping[suffix]
        except KeyError as exc:
            raise DataIntakeError(f"Unsupported input format: {path.name}") from exc

    @staticmethod
    def _make_read_only(path: Path) -> None:
        path.chmod(stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as handle:
            for chunk in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _detect_encoding(path: Path) -> str:
        payload = path.read_bytes()
        if payload.startswith(b"\xef\xbb\xbf"):
            payload.decode("utf-8-sig")
            return "utf-8-sig"
        try:
            payload.decode("utf-8")
            return "utf-8"
        except UnicodeDecodeError:
            payload.decode("latin-1")
            return "latin-1"

    def _file_manifest(
        self,
        path: Path,
        *,
        bundle_root: Path,
        original_name: str,
        file_format: FileFormat,
        parent_source_id: DataSourceId | None = None,
        archive_member: str | None = None,
    ) -> DataFileManifest:
        encoding = None
        if file_format in {FileFormat.CSV, FileFormat.JSON, FileFormat.JSONL}:
            encoding = self._detect_encoding(path)
        return DataFileManifest(
            original_name=original_name,
            stored_relative_path=path.relative_to(bundle_root).as_posix(),
            format=file_format,
            size_bytes=path.stat().st_size,
            sha256=self._sha256(path),
            encoding=encoding,
            parent_source_id=parent_source_id,
            archive_member=archive_member,
        )

    @classmethod
    def _safe_member_path(cls, member_name: str) -> PurePosixPath:
        if "\\" in member_name:
            raise DataIntakeError(f"unsafe ZIP member: {member_name}")
        path = PurePosixPath(member_name)
        if path.is_absolute() or ".." in path.parts or (path.parts and path.parts[0].endswith(":")):
            raise DataIntakeError(f"unsafe ZIP member: {member_name}")
        return path

    def _inspect_zip(
        self,
        path: Path,
        *,
        bundle_root: Path,
        expanded_root: Path,
        parent: DataFileManifest,
        used_table_names: set[str],
    ) -> tuple[list[DataFileManifest], list[TableManifest], list[str]]:
        files: list[DataFileManifest] = []
        tables: list[TableManifest] = []
        warnings: list[str] = []
        names: set[str] = set()

        with zipfile.ZipFile(path) as archive:
            infos = archive.infolist()
            if len(infos) > self.MAX_ZIP_MEMBERS:
                raise DataIntakeError("ZIP member count exceeds safety limit")
            total_bytes = sum(info.file_size for info in infos)
            if total_bytes > self.MAX_ZIP_TOTAL_BYTES:
                raise DataIntakeError("ZIP uncompressed size exceeds safety limit")

            for info in infos:
                member_path = self._safe_member_path(info.filename)
                if info.filename in names:
                    raise DataIntakeError(f"duplicate ZIP member: {info.filename}")
                names.add(info.filename)
                if info.is_dir():
                    continue
                unix_mode = info.external_attr >> 16
                if stat.S_ISLNK(unix_mode):
                    raise DataIntakeError(f"ZIP symbolic links are not allowed: {info.filename}")
                if info.file_size > self.MAX_ZIP_MEMBER_BYTES:
                    raise DataIntakeError(f"ZIP member exceeds size limit: {info.filename}")
                if info.compress_size == 0 and info.file_size > 0:
                    raise DataIntakeError(f"ZIP member has invalid compression metadata: {info.filename}")
                if info.compress_size and info.file_size / info.compress_size > self.MAX_ZIP_COMPRESSION_RATIO:
                    raise DataIntakeError(f"ZIP member exceeds compression-ratio limit: {info.filename}")

                try:
                    file_format = self._format_for_path(Path(member_path.name))
                except DataIntakeError:
                    warnings.append(f"Ignored unsupported ZIP member: {info.filename}")
                    continue
                if file_format is FileFormat.ZIP:
                    warnings.append(f"Ignored nested ZIP member: {info.filename}")
                    continue

                extracted_path = expanded_root.joinpath(*member_path.parts)
                extracted_path.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(info) as source, extracted_path.open("wb") as destination:
                    shutil.copyfileobj(source, destination)
                self._make_read_only(extracted_path)

                member_file = self._file_manifest(
                    extracted_path,
                    bundle_root=bundle_root,
                    original_name=member_path.name,
                    file_format=file_format,
                    parent_source_id=parent.id,
                    archive_member=info.filename,
                )
                files.append(member_file)
                tables.extend(
                    self._inspect_table_file(
                        extracted_path,
                        source_file=member_file,
                        used_table_names=used_table_names,
                    )
                )
        return files, tables, warnings

    def _inspect_table_file(
        self,
        path: Path,
        *,
        source_file: DataFileManifest,
        used_table_names: set[str],
    ) -> list[TableManifest]:
        if source_file.format is FileFormat.CSV:
            return [
                TableManifest(
                    table_name=self._unique_table_name(path.stem, used_table_names),
                    source_file_id=source_file.id,
                    format=source_file.format,
                    archive_member=source_file.archive_member,
                )
            ]
        if source_file.format is FileFormat.JSONL:
            self._validate_jsonl(path, source_file.encoding or "utf-8")
            return [
                TableManifest(
                    table_name=self._unique_table_name(path.stem, used_table_names),
                    source_file_id=source_file.id,
                    format=source_file.format,
                    archive_member=source_file.archive_member,
                )
            ]
        if source_file.format is FileFormat.JSON:
            return self._inspect_json(path, source_file=source_file, used_table_names=used_table_names)
        if source_file.format is FileFormat.XLSX:
            return self._inspect_xlsx(path, source_file=source_file, used_table_names=used_table_names)
        raise DataIntakeError(f"Unsupported table file: {path.name}")

    @staticmethod
    def _validate_jsonl(path: Path, encoding: str) -> None:
        row_count = 0
        with path.open(encoding=encoding) as handle:
            for line_number, line in enumerate(handle, start=1):
                if not line.strip():
                    continue
                try:
                    row = json.loads(line)
                except json.JSONDecodeError as exc:
                    raise DataIntakeError(f"Invalid JSONL at {path.name}:{line_number}") from exc
                if not isinstance(row, dict):
                    raise DataIntakeError(f"JSONL rows must be objects: {path.name}:{line_number}")
                row_count += 1
        if row_count == 0:
            raise DataIntakeError(f"JSONL file has no object rows: {path.name}")

    def _inspect_json(
        self,
        path: Path,
        *,
        source_file: DataFileManifest,
        used_table_names: set[str],
    ) -> list[TableManifest]:
        try:
            payload = json.loads(path.read_text(encoding=source_file.encoding or "utf-8"))
        except json.JSONDecodeError as exc:
            raise DataIntakeError(f"Invalid JSON file: {path.name}") from exc

        table_candidates: list[str]
        if isinstance(payload, list):
            if not all(isinstance(row, dict) for row in payload):
                raise DataIntakeError(f"JSON list rows must be objects: {path.name}")
            table_candidates = [path.stem]
        elif isinstance(payload, dict) and payload and all(isinstance(rows, list) for rows in payload.values()):
            for name, rows in payload.items():
                if not all(isinstance(row, dict) for row in rows):
                    raise DataIntakeError(f"JSON table rows must be objects: {path.name}:{name}")
            table_candidates = list(payload)
        elif isinstance(payload, dict):
            table_candidates = [path.stem]
        else:
            raise DataIntakeError(f"JSON root must be an object or object list: {path.name}")

        object_of_tables = isinstance(payload, dict) and payload and all(isinstance(rows, list) for rows in payload.values())
        return [
            TableManifest(
                table_name=self._unique_table_name(candidate, used_table_names),
                source_file_id=source_file.id,
                format=source_file.format,
                json_key=candidate if object_of_tables else None,
                archive_member=source_file.archive_member,
            )
            for candidate in table_candidates
        ]

    def _inspect_xlsx(
        self,
        path: Path,
        *,
        source_file: DataFileManifest,
        used_table_names: set[str],
    ) -> list[TableManifest]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise DataIntakeError(f"Invalid Excel workbook: {path.name}") from exc
        try:
            return [
                TableManifest(
                    table_name=self._unique_table_name(sheet_name, used_table_names),
                    source_file_id=source_file.id,
                    format=source_file.format,
                    sheet_name=sheet_name,
                    archive_member=source_file.archive_member,
                )
                for sheet_name in workbook.sheetnames
            ]
        finally:
            workbook.close()

    @staticmethod
    def _unique_table_name(candidate: str, used: set[str]) -> str:
        normalized = re.sub(r"[^a-z0-9]+", "_", candidate.strip().lower()).strip("_") or "table"
        table_name = normalized
        suffix = 2
        while table_name in used:
            table_name = f"{normalized}_{suffix}"
            suffix += 1
        used.add(table_name)
        return table_name
