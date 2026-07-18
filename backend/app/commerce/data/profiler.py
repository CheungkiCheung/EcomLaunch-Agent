"""Deterministic schema, quality, and cross-table cardinality profiling."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from enum import StrEnum
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

from app.commerce.data.intake import DataBundleManifest, DataFileManifest, FileFormat, TableManifest
from app.commerce.domain.ids import DatasetId, WorkspaceId


class InferredType(StrEnum):
    EMPTY = "empty"
    BOOLEAN = "boolean"
    INTEGER = "integer"
    DECIMAL = "decimal"
    DATETIME = "datetime"
    STRING = "string"
    JSON = "json"
    MIXED = "mixed"


class JoinCardinality(StrEnum):
    ONE_TO_ONE = "one_to_one"
    ONE_TO_MANY = "one_to_many"
    MANY_TO_ONE = "many_to_one"
    MANY_TO_MANY = "many_to_many"


class ProfileModel(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)


class ColumnProfile(ProfileModel):
    name: str = Field(min_length=1)
    inferred_type: InferredType
    row_count: int = Field(ge=0)
    non_null_count: int = Field(ge=0)
    missing_count: int = Field(ge=0)
    missing_rate: float = Field(ge=0.0, le=1.0)
    unique_count: int = Field(ge=0)
    unique_rate: float = Field(ge=0.0, le=1.0)
    example_values: tuple[str, ...] = Field(default_factory=tuple)
    numeric_min: Decimal | None = None
    numeric_max: Decimal | None = None
    leading_zero_count: int = Field(default=0, ge=0)
    leading_zero_rate: float = Field(default=0.0, ge=0.0, le=1.0)
    is_primary_key_candidate: bool = False
    is_time_candidate: bool = False


class TableProfile(ProfileModel):
    table_name: str = Field(min_length=1)
    row_count: int = Field(ge=0)
    column_count: int = Field(ge=0)
    columns: tuple[ColumnProfile, ...]
    duplicate_row_count: int = Field(ge=0)
    duplicate_row_rate: float = Field(ge=0.0, le=1.0)
    primary_key_candidates: tuple[str, ...] = Field(default_factory=tuple)
    time_candidates: tuple[str, ...] = Field(default_factory=tuple)

    def column(self, name: str) -> ColumnProfile:
        for column in self.columns:
            if column.name == name:
                return column
        raise KeyError(f"Unknown column {self.table_name}.{name}")


class JoinRisk(ProfileModel):
    left_table: str
    left_column: str
    right_table: str
    right_column: str
    cardinality: JoinCardinality
    requires_aggregation: bool
    reason: str


class DatasetProfile(ProfileModel):
    schema_version: str = "1.0"
    dataset_id: DatasetId
    workspace_id: WorkspaceId
    tables: tuple[TableProfile, ...]
    join_risks: tuple[JoinRisk, ...] = Field(default_factory=tuple)

    def table(self, name: str) -> TableProfile:
        for table in self.tables:
            if table.table_name == name:
                return table
        raise KeyError(f"Unknown table {name}")

    def join(self, left_table: str, left_column: str, right_table: str, right_column: str) -> JoinRisk:
        for risk in self.join_risks:
            if (risk.left_table, risk.left_column, risk.right_table, risk.right_column) == (
                left_table,
                left_column,
                right_table,
                right_column,
            ):
                return risk
            if (risk.left_table, risk.left_column, risk.right_table, risk.right_column) == (
                right_table,
                right_column,
                left_table,
                left_column,
            ):
                inverse = {
                    JoinCardinality.ONE_TO_ONE: JoinCardinality.ONE_TO_ONE,
                    JoinCardinality.ONE_TO_MANY: JoinCardinality.MANY_TO_ONE,
                    JoinCardinality.MANY_TO_ONE: JoinCardinality.ONE_TO_MANY,
                    JoinCardinality.MANY_TO_MANY: JoinCardinality.MANY_TO_MANY,
                }[risk.cardinality]
                return risk.model_copy(
                    update={
                        "left_table": left_table,
                        "left_column": left_column,
                        "right_table": right_table,
                        "right_column": right_column,
                        "cardinality": inverse,
                    }
                )
        raise KeyError(f"Unknown join {left_table}.{left_column} -> {right_table}.{right_column}")


class DataProfiler:
    """Profile stored tables without coercing or mutating source values."""

    def __init__(self, *, storage_root: Path):
        self._storage_root = storage_root

    def profile(self, manifest: DataBundleManifest) -> DatasetProfile:
        table_rows: dict[str, list[dict[str, Any]]] = {}
        table_profiles: list[TableProfile] = []

        for table in manifest.tables:
            rows = self.read_rows(manifest, table)
            table_rows[table.table_name] = rows
            table_profiles.append(self._profile_table(table.table_name, rows))

        return DatasetProfile(
            dataset_id=manifest.dataset_id,
            workspace_id=manifest.workspace_id,
            tables=tuple(table_profiles),
            join_risks=self._profile_joins(table_profiles),
        )

    def read_rows(self, manifest: DataBundleManifest, table: TableManifest) -> list[dict[str, Any]]:
        bundle_root = self._storage_root / manifest.storage_relative_path
        files = {file.id: file for file in manifest.files}
        source_file = files[table.source_file_id]
        path = bundle_root / source_file.stored_relative_path
        return self._read_rows(path, source_file=source_file, table=table)

    def _read_rows(
        self,
        path: Path,
        *,
        source_file: DataFileManifest,
        table: TableManifest,
    ) -> list[dict[str, Any]]:
        encoding = source_file.encoding or "utf-8"
        if table.format is FileFormat.CSV:
            with path.open(encoding=encoding, newline="") as handle:
                return [dict(row) for row in csv.DictReader(handle)]
        if table.format is FileFormat.JSONL:
            with path.open(encoding=encoding) as handle:
                return [json.loads(line) for line in handle if line.strip()]
        if table.format is FileFormat.JSON:
            payload = json.loads(path.read_text(encoding=encoding))
            if table.json_key is not None:
                return [dict(row) for row in payload[table.json_key]]
            if isinstance(payload, list):
                return [dict(row) for row in payload]
            return [dict(payload)]
        if table.format is FileFormat.XLSX:
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook[table.sheet_name or workbook.sheetnames[0]]
                iterator = sheet.iter_rows(values_only=True)
                header_row = next(iterator, ())
                headers = tuple(self._header_name(value, index) for index, value in enumerate(header_row, start=1))
                return [dict(zip(headers, row, strict=False)) for row in iterator]
            finally:
                workbook.close()
        raise ValueError(f"Unsupported table format: {table.format}")

    @staticmethod
    def _header_name(value: Any, index: int) -> str:
        text = str(value).strip() if value is not None else ""
        return text or f"column_{index}"

    def _profile_table(self, table_name: str, rows: list[dict[str, Any]]) -> TableProfile:
        column_names: list[str] = []
        seen_columns: set[str] = set()
        for row in rows:
            for column in row:
                if column not in seen_columns:
                    column_names.append(column)
                    seen_columns.add(column)

        columns = tuple(self._profile_column(name, [row.get(name) for row in rows]) for name in column_names)
        row_keys = [tuple(self._canonical(row.get(name)) for name in column_names) for row in rows]
        duplicate_count = sum(count - 1 for count in Counter(row_keys).values() if count > 1)
        row_count = len(rows)
        return TableProfile(
            table_name=table_name,
            row_count=row_count,
            column_count=len(columns),
            columns=columns,
            duplicate_row_count=duplicate_count,
            duplicate_row_rate=duplicate_count / row_count if row_count else 0.0,
            primary_key_candidates=tuple(column.name for column in columns if column.is_primary_key_candidate),
            time_candidates=tuple(column.name for column in columns if column.is_time_candidate),
        )

    def _profile_column(self, name: str, values: list[Any]) -> ColumnProfile:
        row_count = len(values)
        present = [value for value in values if not self._is_missing(value)]
        canonical_values = [self._canonical(value) for value in present]
        unique_values = set(canonical_values)
        inferred_types = {self._infer_value_type(value) for value in present}
        inferred_type = self._combine_types(inferred_types)
        numeric_values = [self._to_decimal(value) for value in present] if inferred_type in {InferredType.INTEGER, InferredType.DECIMAL} else []
        leading_zero_count = sum(isinstance(value, str) and re.fullmatch(r"0\d+", value.strip()) is not None for value in present)
        non_null_count = len(present)
        missing_count = row_count - non_null_count
        unique_count = len(unique_values)
        lower_name = name.lower()
        is_time_candidate = inferred_type is InferredType.DATETIME or any(
            token in lower_name for token in ("date", "time", "timestamp")
        ) or lower_name.endswith("_at")
        is_primary_key = row_count > 0 and missing_count == 0 and unique_count == row_count
        examples = tuple(dict.fromkeys(canonical_values))[:5]
        return ColumnProfile(
            name=name,
            inferred_type=inferred_type,
            row_count=row_count,
            non_null_count=non_null_count,
            missing_count=missing_count,
            missing_rate=missing_count / row_count if row_count else 0.0,
            unique_count=unique_count,
            unique_rate=unique_count / non_null_count if non_null_count else 0.0,
            example_values=examples,
            numeric_min=min(numeric_values) if numeric_values else None,
            numeric_max=max(numeric_values) if numeric_values else None,
            leading_zero_count=leading_zero_count,
            leading_zero_rate=leading_zero_count / non_null_count if non_null_count else 0.0,
            is_primary_key_candidate=is_primary_key,
            is_time_candidate=is_time_candidate,
        )

    @staticmethod
    def _is_missing(value: Any) -> bool:
        return value is None or (isinstance(value, str) and value.strip() == "")

    @staticmethod
    def _canonical(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (dict, list, tuple)):
            return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        return str(value)

    @classmethod
    def _infer_value_type(cls, value: Any) -> InferredType:
        if isinstance(value, bool):
            return InferredType.BOOLEAN
        if isinstance(value, int):
            return InferredType.INTEGER
        if isinstance(value, (float, Decimal)):
            return InferredType.DECIMAL
        if isinstance(value, (dict, list, tuple)):
            return InferredType.JSON
        if not isinstance(value, str):
            return InferredType.STRING

        text = value.strip()
        if re.fullmatch(r"0\d+", text):
            return InferredType.STRING
        if text.lower() in {"true", "false"}:
            return InferredType.BOOLEAN
        if re.fullmatch(r"[-+]?\d+", text):
            return InferredType.INTEGER
        if re.fullmatch(r"[-+]?(?:\d+\.\d*|\d*\.\d+)", text):
            return InferredType.DECIMAL
        try:
            datetime.fromisoformat(text.replace("Z", "+00:00"))
        except ValueError:
            return InferredType.STRING
        return InferredType.DATETIME

    @staticmethod
    def _combine_types(types: set[InferredType]) -> InferredType:
        if not types:
            return InferredType.EMPTY
        if types <= {InferredType.INTEGER, InferredType.DECIMAL}:
            return InferredType.DECIMAL if InferredType.DECIMAL in types else InferredType.INTEGER
        if len(types) == 1:
            return next(iter(types))
        return InferredType.MIXED

    @staticmethod
    def _to_decimal(value: Any) -> Decimal:
        try:
            return Decimal(str(value).strip())
        except InvalidOperation as exc:
            raise ValueError(f"Cannot convert numeric value to Decimal: {value!r}") from exc

    @staticmethod
    def _profile_joins(tables: list[TableProfile]) -> tuple[JoinRisk, ...]:
        risks: list[JoinRisk] = []
        for left_index, left in enumerate(tables):
            left_columns = {column.name: column for column in left.columns}
            for right in tables[left_index + 1 :]:
                right_columns = {column.name: column for column in right.columns}
                for name in sorted(left_columns.keys() & right_columns.keys()):
                    left_column = left_columns[name]
                    right_column = right_columns[name]
                    left_unique = left_column.missing_count == 0 and left_column.unique_count == left_column.row_count
                    right_unique = right_column.missing_count == 0 and right_column.unique_count == right_column.row_count
                    if left_unique and right_unique:
                        cardinality = JoinCardinality.ONE_TO_ONE
                    elif left_unique:
                        cardinality = JoinCardinality.ONE_TO_MANY
                    elif right_unique:
                        cardinality = JoinCardinality.MANY_TO_ONE
                    else:
                        cardinality = JoinCardinality.MANY_TO_MANY
                    risks.append(
                        JoinRisk(
                            left_table=left.table_name,
                            left_column=name,
                            right_table=right.table_name,
                            right_column=name,
                            cardinality=cardinality,
                            requires_aggregation=cardinality is not JoinCardinality.ONE_TO_ONE,
                            reason=f"{name} is {cardinality.value}; aggregate before joining when row multiplication is unintended.",
                        )
                    )
        return tuple(risks)
